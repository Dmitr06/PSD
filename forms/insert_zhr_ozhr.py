from openpyxl import load_workbook
from datetime import *
import os, os.path


def search_content(ws, text, current_row):
    for i in ws.iter_rows(min_row=current_row, max_col=1):  # заполняем концовку 0 раздела
        if i[0].value == text:
            current_row = i[0].row + 1
    return current_row


def insert_zhr_ozhr(akt_vh, def_db, road_programm, road_db, tube, km_start, km_finish, dy_tube):
    def table_main(ws, ws_style, table, current_row, coord, height=13.50):
        if table == table_0:
            for i in table:
                ws.merge_cells(i.format(current_row))
            j = 30
        else:
            j = len(table) + 1
        ws.row_dimensions[current_row].height = height
        for index in range(1, j):
            ws.cell(current_row, index)._style = ws_style.cell(*coord)._style

    wb_zhr = load_workbook(road_programm + road_to_excel)
    ws_tittle = wb_zhr['L_tittle']
    ws_tittle['R1'], ws_tittle['R2'] = km_start + ' км', km_finish + ' км'
    ws_tittle['R3'], ws_tittle['R4'] = def_db[0]['date'][0].strftime('%d.%m.%Y') + ' г.', def_db[-1]['date'][2].strftime('%d.%m.%Y') + ' г.'
    ws = wb_zhr['L1']
    ws_l2 = wb_zhr['L2']
    otv = []
    contr = []
    sk = []
    lkk = []
    section_3_content = {}
    table_0 = ('A{0}:I{0}', 'J{0}:N{0}', 'O{0}:Y{0}', 'Z{0}:AC{0}')  # table of number 0
    table_1 = ('A', 'B', 'C', 'D', 'E')  # table of section 1
    table_2 = ('A', 'C', 'D', 'E', 'F')  # table of section 2
    table_3 = ('A', 'C', 'D')  # table of section 3
    table_6 = ('A', 'C')  # table of section 6
    # -----------------------SECTION 0---------------
    ws['A1'] = def_db[0]['otv'][2]
    ws['A3'] = 'по оъекту: Устранение дефектов на секциях %s, %s-%s км, Ду %s мм.' % (
        tube, km_start, km_finish, dy_tube)
    for sec in def_db:
        for person, list_persons in zip(('otv', 'contr', 'sk', 'lkk'), (otv, contr, sk, lkk)):
            if sec[person] not in list_persons:
                list_persons.append(sec[person])
    ws['A30'] = contr[0][0]
    ws['J30'] = contr[0][1] + ' ' + contr[0][2]
    if len(contr) > 1:
        ws['A31'] = contr[1][0]
        ws['J31'] = contr[1][1] + ' ' + contr[1][2]
        ws['O31'] = ws['O30'].value
    ws['C69'] = otv[0][2]
    current_row = 79  # current select row
    for otv_man in otv:  # заполняем застройщмка в строке 78
        ws.insert_rows(current_row)
        table_main(ws, ws_l2, table_0, current_row, (1, 1), 55)
        ws['A' + str(current_row)] = otv_man[0]
        ws['J' + str(current_row)] = otv_man[1] + otv_man[2]
        ws['O' + str(current_row)] = ws['O30'].value
        current_row += 1

    current_row = search_content(ws, 'Общие сведения об объекте капитального строительства',
                                 current_row)  # заполняем концовку 0 раздела
    ws.merge_cells('A{0}:AC{1}'.format(current_row, current_row + 1))
    ws['A' + str(current_row)] = 'Устранение дефектов на секциях %s, %s-%s км, Ду %s мм.' % (
        tube, km_start, km_finish, dy_tube)
    current_row = search_content(ws, 'Начало строительства, реконструкции, капитального ремонта  объекта', current_row)
    ws['K' + str(current_row)] = def_db[0]['date'][0].strftime('%d.%m.%Y') + ' г.'
    ws['N' + str(current_row + 3)] = def_db[-1]['date'][2].strftime('%d.%m.%Y') + ' г.'
    ws.print_area = 'A1:AC' + str(current_row + 36)
    # ---------------SECTION 1---------------
    ws1 = wb_zhr['Part 1']
    current_row = 7
    for otv_man in otv:  # заносим в таблицу ответственного за проведение работ
        ws1.insert_rows(current_row)
        table_main(ws1, ws_l2, table_1, current_row, (1, 2), 54)
        ws1['A' + str(current_row)] = otv_man[2]
        ws1['B' + str(current_row)] = f'{otv_man[1]} {otv_man[2]} {otv_man[0]}'
        work_start = None
        for i in def_db:
            if otv_man is i['otv']:
                if work_start == None:
                    work_start = i['date'][0]
                work_stop = i['date'][2]
        ws1['C' + str(current_row)] = work_start.strftime(
            '%d.%m.%Y') + ' г. ' + 'Основные работы по устранению дефектов.'
        ws1['D' + str(current_row)] = work_stop.strftime('%d.%m.%Y') + ' г.'
        ws1['E' + str(current_row)] = ws1['B' + str(current_row)].value
        current_row += 1
    for otv_man in contr:  # заносим в таблицу ответственного за контроль
        ws1.insert_rows(current_row)
        table_main(ws1, ws_l2, table_1, current_row, (1, 2), 54)
        ws1['A' + str(current_row)] = otv_man[2]
        ws1['B' + str(current_row)] = f'{otv_man[1]} {otv_man[2]} {otv_man[0]}'
        work_start = None
        for i in def_db:
            if otv_man == i['contr']:
                if work_start == None:
                    work_start = i['date'][0]
                work_stop = i['date'][2]
        ws1['C' + str(current_row)] = work_start.strftime('%d.%m.%Y') + ' г. ' + 'Контроль  за подрядной организа-цией.'
        ws1['D' + str(current_row)] = work_stop.strftime('%d.%m.%Y') + ' г.'
        ws1['E' + str(current_row)] = ws1['B' + str(current_row)].value
        current_row += 1
    for otv_man in lkk:  # заносим в таблицу дефектоскопистов
        ws1.insert_rows(current_row)
        table_main(ws1, ws_l2, table_1, current_row, (1, 2), 54)
        ws1['A' + str(current_row)] = otv_man[2]
        ws1['B' + str(current_row)] = f'{otv_man[1]} {otv_man[2]} {otv_man[0]}'
        work_start = None
        for i in def_db:
            if otv_man == i['lkk']:
                if work_start == None:
                    work_start = i['date'][0]
                work_stop = i['date'][2]
        ws1['C' + str(current_row)] = work_start.strftime('%d.%m.%Y') + ' г. ' + 'Проведение НК сварных соединений.'
        ws1['D' + str(current_row)] = work_stop.strftime('%d.%m.%Y') + ' г.'
        ws1['E' + str(current_row)] = ws1['B' + str(current_row)].value
        current_row += 1
    if current_row > 17:
        ws1.print_area = 'A1:F28'
    # ---------------SECTION 2---------------
    ws2 = wb_zhr['Part 2']
    current_row = 5
    for i in range(1, 7):  # insert in table info
        current_row += 1
        table_main(ws2, ws_l2, table_2, current_row, (1, 2), height=55)
        ws2['C' + str(current_row)] = ' '.join((otv[0][1], otv[0][2], otv[0][0]))
        ws2['D' + str(current_row)] = work_stop.strftime('%d.%m.%Y') + ' г.'
        ws2['E' + str(current_row)] = ' '.join((contr[0][1], contr[0][2], contr[0][0]))
    ws2['C' + str(current_row)] = ' '.join((lkk[0][1], lkk[0][2], lkk[0][0]))

    # ---------------SECTION 3---------------
    ws3 = wb_zhr['Part 3']
    current_row = 5
    for sec in def_db:  # creat the list with day,work,man
        ground_work = int(sec['rand_value'][2] * sec['rand_value'][3] * (
                    sec['dist'][4] - sec['dist'][3]) * 1.4)  # calculation ground work
        for i, date in enumerate(sec['date']):
            date_temp = date.strftime('%d.%m.%Y') + ' г.'  # creat key for dict
            otv_temp = ' '.join((sec['otv'][1], sec['otv'][2], sec['otv'][0]))
            lkk_temp = ' '.join((sec['lkk'][1], sec['lkk'][2], sec['lkk'][0]))
            if 'grnd_maker' in sec:
                grnd_maker_temp = ' '.join((sec['grnd_maker'][1], sec['grnd_maker'][2], sec['grnd_maker'][0]))
            else:
                grnd_maker_temp = ' '.join((sec['otv'][1], sec['otv'][2], sec['otv'][0]))
            if date_temp not in section_3_content:  # insert new days
                section_3_content[date_temp] = []
            if i == 0:
                section_3_content[date_temp].append(('Вскрытие дефектного участка трубопровода на секции № %s - %s м3.'
                                                     % (sec['sec'], ground_work), grnd_maker_temp))
                section_3_content[date_temp].append(('Очистка трубопровода от изоляции на секции № %s - %s п.м.'
                                                     % (sec['sec'], round((sec['dist'][2] - sec['dist'][1]), 1)),
                                                     otv_temp))
                section_3_content[date_temp].append(('Проведение ДДК секции № %s.' % sec['sec'], lkk_temp))
                type_repair = True
                for j in sec['defect'].values():
                    if 'П1' in j['type']:
                        type_repair = False
                        section_3_content[date_temp].append(
                            ('Установка муфты П1 на секции № %s - 1 шт.' % (sec['sec']), otv_temp))
                        section_3_content[date_temp].append(('Дефектоскопия сварных стыков - 2 шт.', lkk_temp))
                        break
                    elif 'уфт' in j['type']:
                        type_repair = False
                        section_3_content[date_temp].append(
                            ('Установка муфты П2 на секции № %s - 1 шт.' % (sec['sec']), otv_temp))
                        section_3_content[date_temp].append(('Дефектоскопия сварных стыков - 10 шт.', lkk_temp))
                        break
                if type_repair == True:  # type of repair/mufts
                    section_3_content[date_temp].append(
                        ('Шлифовка дефектов на секции № %s - %s дефектов.' % (sec['sec'], len(j)), otv_temp))
                    section_3_content[date_temp].append(
                        ('Дефектоскопия отремонтированных дефектов - %s дефектов.' % len(j), lkk_temp))
            if i == 1:
                section_3_content[date_temp].append(('Изоляция отремонтированного участка секции № %s - %s п.м.' % (
                sec['sec'], round((sec['dist'][2] - sec['dist'][1] + 0.45), 1)), otv_temp))
            if i == 2:
                section_3_content[date_temp].append(('Определение адгезии изоляционного покрытия секции № %s.' % sec[
                    'sec'], 'Мастер РУ №2 ЦРС "Рязань" Лазарев А.В.'))
                section_3_content[date_temp].append(
                    ('Обратная засыпка  дефектного участка трубопровода на секции № %s - %s м3.'
                     % (sec['sec'], ground_work), grnd_maker_temp))
    for days in section_3_content.keys():  # insert in excel from list section_3_content
        for day in section_3_content[days]:  # for everyday
            current_row += 1
            table_main(ws3, ws_l2, table_3, current_row, (1, 2), 35)
            ws3['A' + str(current_row)] = days
            ws3['B' + str(current_row)] = day[0]
            ws3['C' + str(current_row)] = day[1]
    if (current_row - 5) % 18 != 0:
        page_end = 18 - (current_row - 5) + (current_row - 5) // 18 * 18
        for i in range(current_row, current_row + page_end):
            current_row += 1
            table_main(ws3, ws_l2, table_3, current_row, (1, 2), 35)
    ws3.print_area = 'A1:C%s' % current_row
    # ---------------SECTION 6---------------
    ws6 = wb_zhr['Part 6']
    current_row = 5
    index_temp = 3
    for numb_akt, akt in enumerate(akt_vh):  # acts of incoming control
        current_row += 1
        temp_otv = ' '.join((akt['otv'][1], akt['otv'][2], akt['otv'][0]))
        temp_contr = ' '.join((akt['contr'][1], akt['contr'][2], akt['contr'][0]))
        temp_sk = ' '.join((akt['sk'][1], akt['sk'][2], akt['sk'][0]))
        ws6.insert_rows(current_row)
        table_main(ws6, ws_l2, table_6, current_row, (1, 2), 55)
        ws6['A' + str(current_row)] = 'Акт входного контроля на %s №%s' % (akt['full_name'], numb_akt + 3)
        ws6['B' + str(current_row)] = '%s г. \n%s \n%s \n%s' % (akt['date'], temp_otv, temp_contr, temp_sk)

    for numb_defect, akt in enumerate(def_db):
        current_row += 1
        temp_otv = ' '.join((akt['otv'][1], akt['otv'][2], akt['otv'][0]))
        temp_contr = ' '.join((akt['contr'][1], akt['contr'][2], akt['contr'][0]))
        temp_sk = ' '.join((akt['sk'][1], akt['sk'][2], akt['sk'][0]))
        if 'grnd_maker' in akt:
            temp_grnd_maker = ' '.join((akt['grnd_maker'][1], akt['grnd_maker'][2], akt['grnd_maker'][0]))
            temp_grnd_contr = ' '.join((akt['grnd_contr'][1], akt['grnd_contr'][2], akt['grnd_contr'][0]))
        else:
            temp_grnd_maker = ' '.join((akt['otv'][1], akt['otv'][2], akt['otv'][0]))
            temp_grnd_contr = ' '.join((akt['contr'][1], akt['contr'][2], akt['contr'][0]))
        temp_date = akt['date'][2].strftime('%d.%m.%Y')
        ws6.insert_rows(current_row)
        table_main(ws6, ws_l2, table_6, current_row, (1, 2), 55)
        ws6['A' + str(current_row)] = 'Акт о выборочном ремонте дефектов на секции %s - №%s' % (
        akt['sec'], numb_defect + 1)
        ws6['B' + str(current_row)] = '%s г. \n%s \n%s \n%s' % (temp_date, temp_otv, temp_contr, temp_sk)
        current_row += 1
        ws6.insert_rows(current_row)
        table_main(ws6, ws_l2, table_6, current_row, (1, 2), 55)
        ws6['A' + str(current_row)] = 'Акты определения адгезии защитных покрытий секции %s - №%s' % (
        akt['sec'], numb_defect + 1)
        ws6['B' + str(current_row)] = '%s г. \n%s,\n%s,\n%s' % (temp_date, temp_otv, temp_contr, temp_sk)
        current_row += 1
        index_temp += 1
        ws6.insert_rows(current_row)
        table_main(ws6, ws_l2, table_6, current_row, (1, 2), 55)
        ws6['A' + str(current_row)] = 'Акт на засыпку (обваловку) уложенного трубопровода - №%s' % (numb_defect + 1)
        ws6['B' + str(current_row)] = '%s г. \n%s \n%s \n%s' % (temp_date, temp_grnd_maker, temp_grnd_contr, temp_sk)
        current_row += 1
        ws6.insert_rows(current_row)
        table_main(ws6, ws_l2, table_6, current_row, (1, 2), 55)
        ws6['A' + str(
            current_row)] = 'Акт на фактически выполненный объём работ при устранении дефектов  на секции %s - №%s' % (
        akt['sec'], numb_defect + 1)
        ws6['B' + str(current_row)] = '%s г. \n%s \n%s \n%s' % (temp_date, temp_otv, temp_contr, temp_sk)
    if (current_row - 5) % 11 != 0:
        page_end = current_row - (current_row - 5) % 11 + 11
        for i in range(current_row, page_end):
            current_row += 1
            table_main(ws6, ws_l2, table_6, current_row, (1, 2), 55)
    ws6.print_area = 'A1:B%s' % current_row
    # -----------------END------------------
    ws7 = wb_zhr['Part 7']
    ws7['E46'] = '%s _____________ %s г.' % (temp_otv, temp_date)
    del (wb_zhr['L2'])
    road_db = os.path.normpath(road_db)
    road_db = os.path.join(road_db, 'Журналы')
    if not os.path.exists(road_db):
        os.makedirs(road_db)
    road_db = os.path.join(road_db, '3. ОЖР.xlsx')
    wb_zhr.save(road_db)


road_to_excel = '/Excell/zhr_ozhr.xlsx'
if __name__ == '__main__':
    road_to_excel = '../Excell/zhr_ozhr.xlsx'
    road_db = os.getcwd() + '\..\..\Тест'
    road_programm = ''
    tube, km_start, km_finish, dy_tube = 'МНПП "Рязань-Тула-Орел" отвод на новомосковскую НБ, ДТ', '12', '13', '530'
    def_db = [{'sec': '12', 'date': [date(2018, 12, 12), date(2018, 12, 13), date(2018, 12, 14)],
               'dist': [10.1, 9.11, 11.09, 5.44, 14.76], 'km': '1', 'dl_muft': 1.0,
               'rand_value': [0.48533824224793676, 3.6729518931973657, 2.861346564462221, 2.705419511460805],
               'defect': {'1': {'sec': '12', 'dist': '123', 'lab': 'Риска с ППОШ с расслоением', 'dl': '5', 'sh': '6',
                                'gl': '7', 'type': 'Муфта П1'},
                          '2': {'sec': '12', 'dist': '123', 'lab': 'Риска с ППОШ с расслоением и ВНП', 'dl': '234',
                                'sh': '234', 'gl': '65', 'type': 'Муфта П2'},
                          '3': {'sec': '12', 'dist': '534536', 'lab': 'Риска с ППОШ с расслоением и ВНП', 'dl': '45534',
                                'sh': '345345', 'gl': '3535', 'type': 'Шлифовка'}},
               'otv': ('Афанасьев А.Б.', 'Начальник РУ№1', 'ЦРС "Рязань"'),
               'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'),
               'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег"'),
               'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
              {'sec': '13', 'date': [date(2018, 12, 13), date(2018, 12, 13), date(2018, 12, 14)],
               'dist': [20.0, 18.54, 21.46, 15.55, 24.45], 'km': '2', 'dl_muft': 2.0,
               'rand_value': [0.4581174447339836, 2.989229650534175, 2.9019640191285303, 2.8115670668469814],
               'defect': {
                   '1': {'sec': '13', 'dist': '12313', 'lab': 'Потеря метала', 'dl': '12', 'sh': '132', 'gl': '123',
                         'type': 'Шлифовка'},
                   '2': {'sec': '13', 'dist': '2344234', 'lab': 'Потеря метала', 'dl': '34', 'sh': '4343', 'gl': '434',
                         'type': 'Шлифовка'}}, 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'),
               'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('', '', ''),
               'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
              {'sec': '14', 'date': [date(2018, 12, 14), date(2018, 12, 15), date(2018, 12, 16)],
               'dist': [30.0, 27.94, 32.06, 27.74, 32.26], 'km': '3', 'dl_muft': 3.0,
               'rand_value': [0.5619183603415756, 3.19813104261683, 3.0060601832822327, 2.9168215194726352], 'defect': {
                  '1': {'sec': '14', 'dist': '1131213,32', 'lab': 'Риска', 'dl': '1233', 'sh': '12331', 'gl': '2133',
                        'type': 'Муфта П1'},
                  '2': {'sec': '14', 'dist': '321313,21', 'lab': 'Риска', 'dl': '23', 'sh': '234', 'gl': '2342',
                        'type': 'Муфта П1'}}, 'otv': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'),
               'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'),
               'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег'),
               'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
              {'sec': '15', 'date': [date(2018, 12, 14), date(2018, 12, 14), date(2018, 12, 15)],
               'dist': [160.0, 157.71, 162.29, 154.79, 165.21], 'km': '16', 'dl_muft': 4.0,
               'rand_value': [0.2872474356253941, 2.915977705201697, 2.9123558666117617, 2.703436382412381], 'defect': {
                  '1': {'sec': '15', 'dist': '123131,12', 'lab': 'Риска', 'dl': '123', 'sh': '123', 'gl': '13',
                        'type': 'Муфта П2'},
                  '2': {'sec': '15', 'dist': '1231', 'lab': 'Риска', 'dl': '3213', 'sh': '123', 'gl': '312',
                        'type': 'Муфта П2'},
                  '3': {'sec': '15', 'dist': '432423', 'lab': 'Риска', 'dl': '13', 'sh': '123', 'gl': '13',
                        'type': 'Муфта П2'}}, 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'),
               'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('', '', ''),
               'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')}]
    akt_vh = [
        {'name': 'Абразив', 'full_name': 'Абразив 123', 'marker': 'кг', 'TU': 'ТУ5', 'param': '-', 'date': '02.03.2019',
         'doc': 'Паспорт 34234', 'kol': '3', 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'),
         'contr': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'),
         'kk': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег')},
        {'name': 'Полилен-ПЛ', 'full_name': ' материал Полилен-ПЛ', 'marker': 'рул.', 'TU': 'ТУ3',
         'param': 'длина= мм, ширина= мм', 'date': '01.12.2019', 'doc': 'Паспорт 14', 'kol': '1',
         'otv': ('Афанасьев А.Б.', 'Начальник РУ№1', 'ЦРС "Рязань"'),
         'contr': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'),
         'kk': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег')},
        {'name': 'Муфта П1 530', 'full_name': 'Муфта КМТ 530', 'marker': 'шт.', 'TU': 'ТУ8',
         'param': 'диаметр= мм, длина= мм, толщина стенки= м', 'date': '05.12.2019', 'doc': 'Паспорт 12', 'kol': '3534',
         'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'),
         'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'kk': ('', '', ''), 'sk': ('', '', '')},
        {'name': 'Герметик', 'full_name': 'состав полмерный Герметик', 'marker': 'компл.', 'TU': 'ТУ6', 'param': '-',
         'date': '07.12.2019', 'doc': 'Паспорт 645', 'kol': '3',
         'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'),
         'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'),
         'kk': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'),
         'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег')},
        {'name': 'Полилен-ПЛ', 'full_name': ' материал Полилен-ПЛ', 'marker': 'рул.', 'TU': 'ТУ3',
         'param': 'длина= мм, ширина= мм', 'date': '01.12.2019', 'doc': 'Паспорт 14', 'kol': '1',
         'otv': ('Афанасьев А.Б.', 'Начальник РУ№1', 'ЦРС "Рязань"'),
         'contr': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'),
         'kk': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег')},
        {'name': 'Муфта П1 530', 'full_name': 'Муфта КМТ 530', 'marker': 'шт.', 'TU': 'ТУ8',
         'param': 'диаметр= мм, длина= мм, толщина стенки= м', 'date': '05.12.2019', 'doc': 'Паспорт 12', 'kol': '3534',
         'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'),
         'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'kk': ('', '', ''), 'sk': ('', '', '')},
        {'name': 'Герметик', 'full_name': 'состав полмерный Герметик', 'marker': 'компл.', 'TU': 'ТУ6', 'param': '-',
         'date': '07.12.2019', 'doc': 'Паспорт -', 'kol': '3',
         'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'),
         'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'),
         'kk': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'),
         'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег')}]
    insert_zhr_ozhr(akt_vh, def_db, road_programm, road_db, tube, km_start, km_finish, dy_tube)
