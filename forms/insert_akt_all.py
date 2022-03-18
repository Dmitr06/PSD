from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import *


def insert_akt_all(def_db, road_programm, road_db, tube, km_start, km_finish, dy_tube):
    index_muft = 1
    for i, x in enumerate(def_db):
        wb = load_workbook(road_programm + road_to_excel)
        # --------------начинаем заполнение акта адгезии
        ws1 = wb['Адгезия']
        ws1['O11'] = i + 1
        ws1['W14'] = x['date'][2].strftime('%d.%m.%y') + ' г.'
        ws1['J16'] = tube
        ws1['K22'] = dy_tube + ' мм.'
        ws1['D29'] = x['km'] + ' км.'
        ws1['H45'] = ' '.join((x['otv'][1], x['otv'][2], x['otv'][0]))
        # --------------начинаем заполнение акта о выборочном ремонте
        ws2 = wb['Выборочный ремонт']
        ws2['AB5'] = x['date'][1].strftime('%d.%m.%Y') + ' г.'
        ws2['P9'] = i + 1
        ws2['S10'] = ws2['S10'].value.replace('SEC', x['sec'])
        ws2['A11'] = ws2['A11'].value.replace('DEFEC', ', '.join(x['defect'].keys()))
        ws2['L17'] = tube
        ws2['L18'] = x['km'] + ' км.'
        for rows_tabl in range(26, 25 + len(x['defect'])):
            ws2.insert_rows(rows_tabl)
            for coord in (
                    'B{0}:D{0}', 'E{0}:G{0}', 'H{0}:J{0}', 'K{0}:O{0}', 'P{0}:Q{0}', 'R{0}:S{0}', 'T{0}:U{0}',
                    'T{0}:U{0}',
                    'V{0}:X{0}', 'Y{0}:AA{0}'):
                ws2.merge_cells(coord.format(rows_tabl))
            for cell_tab in range(1, 28):
                ws2.cell(rows_tabl, cell_tab)._style = ws2.cell(25, cell_tab)._style
                ws2.row_dimensions[rows_tabl].height = ws2.row_dimensions[25].height
        current_row = 25
        index_def = 1
        for num_def, defect in x['defect'].items():
            num_row = str(current_row)
            ws2['A' + num_row] = index_def
            ws2['B' + num_row] = x['sec']
            ws2['E' + num_row] = num_def
            ws2['H' + num_row] = 'Дист.\n' + defect['dist']
            ws2['K' + num_row] = defect['lab']
            ws2['P' + num_row] = defect['dl']
            ws2['R' + num_row] = defect['sh']
            ws2['T' + num_row] = defect['gl']
            ws2['V' + num_row] = defect['type']
            ws2['Y' + num_row] = ws2['AB5'].value
            index_def += 1
            current_row += 1
        ws2.merge_cells('A{0}:AA{0}'.format(current_row))
        ws2.row_dimensions[current_row].height = 30
        ws2['A' + str(current_row)].font = Font(size=11)
        current_row += 1
        for (worker, temp_rows) in zip(('otv', 'contr', 'sk'), (current_row, current_row + 6, current_row + 12)):
            ws2.row_dimensions[temp_rows].height = 10
            ws2['P' + str(temp_rows + 1)] = x[worker][2]
            ws2.row_dimensions[temp_rows + 2].height = 12
            ws2.row_dimensions[temp_rows + 3].height = 12
            temp_rows += 4
            ws2['B' + str(temp_rows)] = x[worker][1]
            ws2['J' + str(temp_rows)] = x[worker][0]
        ws2.print_area = 'A1:AB' + str(temp_rows + 1)
        # --------------Разрешение на изоляцию
        ws3 = wb['Изоляция']
        ws3['D6'] = x['otv'][2]
        ws3['Q4'] = 'Устранение дефектов на секциях %s, %s-%s км, Ду %s мм.' % (tube, km_start, km_finish, dy_tube)
        ws3['AB13'] = x['date'][1].strftime('%d.%m.%Y') + ' г.'
        ws3['O16'] = ' '.join((x['sk'][1], x['sk'][2], x['sk'][0]))
        ws3['H17'] = ' '.join((x['otv'][1], x['otv'][2], x['otv'][0]))
        ws3['N18'] = ' '.join((x['lkk'][1], x['lkk'][2], x['lkk'][0]))
        ws3['M19'] = tube
        ws3['G22'] = 'Дист. %s м/%s км' % (x['dist'][1], x['km'])
        ws3['S22'] = 'Дист. %s м/%s км' % (x['dist'][2], x['km'])
        # --------------Укладка тп
        ws5 = wb['Укладка тп']
        ws5['A1'] = x['grnd_maker'][2]
        ws5['A3'] = 'Устранение дефектов на секциях %s, %s-%s км, Ду %s мм.' % (tube, km_start, km_finish, dy_tube)
        ws5['P9'] = x['date'][2].strftime('%d.%m.%Y') + ' г.'
        ws5['K23'] = 'Дист. %s м/%s км' % (x['dist'][3], x['km'])
        ws5['V23'] = 'Дист. %s м/%s км' % (x['dist'][4], x['km'])
        ws5['AA22'] = str(round(x['dist'][4] - x['dist'][3], 1)) + ' м.п.'
        ws5['O13'] = ' '.join((x['grnd_contr'][1],
                               x['grnd_contr'][2],
                               x['grnd_contr'][0]))  # поменять первую букву регистра с большой на маленькую
        ws5['O20'] = ' '.join((x['grnd_maker'][1],
                               x['grnd_maker'][2],
                               x['grnd_maker'][0]))
        ws5['O16'] = ' '.join((x['sk'][1], x['sk'][2], x['sk'][0]))
        # --------------Акт объемов
        ws6 = wb['Объемы']
        ws6['O9'] = i + 1
        ws6['V6'] = x['date'][2].strftime('%d.%m.%Y') + ' г.'
        ws6['L15'] = tube
        ws6['L16'] = x['km'] + ' км'
        ws6['A18'] = ws6['A18'].value % (x['date'][0].strftime('%d.%m.%Y'), x['date'][2].strftime('%d.%m.%Y'), tube)
        for cell_table in ('B25', 'B27', 'B32'):
            ws6[cell_table] = ws6[cell_table].value % (x['sec'])
        current_row = 34
        for (worker, temp_rows) in zip(('otv', 'contr', 'sk'), (current_row, current_row + 6, current_row + 12)):
            ws6['O' + str(temp_rows)] = x[worker][2]
            temp_rows += 3
            ws6['A' + str(temp_rows)] = x[worker][1]
            ws6['I' + str(temp_rows)] = x[worker][0]
        ws6['R25'] = x['date'][0].strftime('%d.%m.%Y') + ' г.'
        ws6['R30'] = x['date'][1].strftime('%d.%m.%Y') + ' г.'
        ws6['R31'] = x['date'][2].strftime('%d.%m.%Y') + ' г.'
        ws6['V25'] = x['grnd_maker'][2]
        ws6['V26'] = x['otv'][2]
        ws6['P25'] = x['rand_value'][2] * x['rand_value'][3] * (x['dist'][4] - x['dist'][3]) * 1.4  # ПРОВЕРИТЬ
        ws6['P26'] = x['dist'][2] - x['dist'][1]
        type_repair = True
        for defect in x['defect'].values():
            if 'уфта' in defect['type']:
                type_repair = False
                ws6['B28'] = 'Установка ремонтной конструкции %s на секции № %s.' % (defect['type'][-2:], x['sec'])
                if '1' not in defect['type']:
                    ws6['P29'] = '10'
        if type_repair == True:
            ws6['B28'] = 'Устранение дефектов методом шлифовки на секции № %s.' % (x['sec'])
            ws6['N28'] = 'Деф.'
            ws6['P28'] = len(x['defect'])
            ws6['B29'] = 'НК отремонтированных дефектов.'
            ws6['N29'] = ws6['N28'].value
            ws6['P29'] = ws6['P28'].value
        # --------------Схемы
        ws7 = wb['Схема']
        temp_index_muft = index_muft
        for defect in x['defect'].values():
            if 'уфта' in defect['type']:
                if '1' in defect['type']:
                    ws7['A7'] = ws7['A7'].value % ('П1', x['sec'])
                    ws7['C36'], ws7['C37'] = '1', '2'
                else:
                    ws7['A7'] = ws7['A7'].value % (defect['type'][-3:], x['sec'])
                ws7['V4'] = index_muft
                ws7['F10'] = ' '.join((tube, x['km'], 'км.'))
                ws7['M13'] = x['otv'][2]
                ws7['F52'] = ' '.join((x['otv'][1], x['otv'][2], x['otv'][0]))
                index_muft += 1
                break
        if index_muft == temp_index_muft:
            del (wb['Схема'])

        wb.save(road_db.rpartition('/')[0] + '/%s. Секция %s.xlsx' % (i + 1, x['sec']))


road_to_excel = '/Excell/akt_all.xlsx'
ws = ('Адгезия', 'Выборочный ремонт', 'Изоляция', 'Рекультивация', 'Укладка тп', 'Объемы', 'Схема')
if __name__ == '__main__':
    import os
    road_to_excel = '../Excell/akt_all.xlsx'
    road = os.getcwd()
    road_db=road.replace('\\','/')[:-9]+'Тест/'
    print(road_db)
    road_programm = ''
    tube, km_start, km_finish, dy_tube = 'МНПП "Рязань-Тула-Орел" отвод на новомосковскую НБ, ДТ', '12', '13', '530'
    def_db = [{'sec': '12', 'date': [date(2018, 12, 12), date(2018, 12, 13), date(2018, 12, 14)],
               'dist': [10.1, 9.11, 11.09, 5.44, 14.76], 'km': '1', 'dl_muft': 1.0,
               'rand_value': [0.48533824224793676, 3.6729518931973657, 2.861346564462221, 2.705419511460805],
               'defect': {'1': {'sec': '12', 'dist': '123', 'lab': 'Риска с ППОШ с расслоением', 'dl': '5', 'sh': '6',
                                'gl': '7', 'type': 'Муфта П1'},
                          '2': {'sec': '12', 'dist': '123', 'lab': 'Риска с ППОШ с расслоением и ВНП', 'dl': '234',
                                'sh': '234', 'gl': '65', 'type': 'Муфта П2'},
                          '3': {'sec': '12', 'dist': '534536', 'lab': 'Риска с ППОШ с расслоением и ВНП', 'dl': '45534',
                                'sh': '345345', 'gl': '3535', 'type': 'Шлифовка'}},
               'otv': ('Афанасьев А.Б.', 'Начальник РУ№1', 'ЦРС "Рязань"'),
               'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'),
               'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег'),
               'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
              {'sec': '13', 'date': [date(2018, 12, 13), date(2018, 12, 13), date(2018, 12, 14)],
               'dist': [20.0, 18.54, 21.46, 15.55, 24.45], 'km': '2', 'dl_muft': 2.0,
               'rand_value': [0.4581174447339836, 2.989229650534175, 2.9019640191285303, 2.8115670668469814],
               'defect': {
                   '1': {'sec': '13', 'dist': '12313', 'lab': 'Потеря метала', 'dl': '12', 'sh': '132', 'gl': '123',
                         'type': 'Шлифовка'},
                   '2': {'sec': '13', 'dist': '2344234', 'lab': 'Потеря метала', 'dl': '34', 'sh': '4343', 'gl': '434',
                         'type': 'Шлифовка'}}, 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'),
               'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('', '', ''),
               'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
              {'sec': '14', 'date': [date(2018, 12, 14), date(2018, 12, 15), date(2018, 12, 16)],
               'dist': [30.0, 27.94, 32.06, 27.74, 32.26], 'km': '3', 'dl_muft': 3.0,
               'rand_value': [0.5619183603415756, 3.19813104261683, 3.0060601832822327, 2.9168215194726352], 'defect': {
                  '1': {'sec': '14', 'dist': '1131213,32', 'lab': 'Риска', 'dl': '1233', 'sh': '12331', 'gl': '2133',
                        'type': 'Муфта П1'},
                  '2': {'sec': '14', 'dist': '321313,21', 'lab': 'Риска', 'dl': '23', 'sh': '234', 'gl': '2342',
                        'type': 'Муфта П1'}}, 'otv': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'),
               'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'),
               'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег'),
               'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
              {'sec': '15', 'date': [date(2018, 12, 14), date(2018, 12, 14), date(2018, 12, 15)],
               'dist': [160.0, 157.71, 162.29, 154.79, 165.21], 'km': '16', 'dl_muft': 4.0,
               'rand_value': [0.2872474356253941, 2.915977705201697, 2.9123558666117617, 2.703436382412381], 'defect': {
                  '1': {'sec': '15', 'dist': '123131,12', 'lab': 'Риска', 'dl': '123', 'sh': '123', 'gl': '13',
                        'type': 'Муфта П2'},
                  '2': {'sec': '15', 'dist': '1231', 'lab': 'Риска', 'dl': '3213', 'sh': '123', 'gl': '312',
                        'type': 'Муфта П2'},
                  '3': {'sec': '15', 'dist': '432423', 'lab': 'Риска', 'dl': '13', 'sh': '123', 'gl': '13',
                        'type': 'Муфта П2'}}, 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'),
               'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('', '', ''),
               'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')}]
    insert_akt_all(def_db, road_programm, road_db, tube, km_start, km_finish, dy_tube)
