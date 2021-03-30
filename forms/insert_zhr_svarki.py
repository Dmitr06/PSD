from openpyxl import load_workbook
import random
from datetime import *
import os, os.path

def create_row (ws,ws2,current_row,coord,height=13.50):
    ws.insert_rows(current_row)
    ws.row_dimensions[current_row].height = height
    for index in range(1,47):
        ws.cell(current_row,index)._style = ws2.cell(*coord)._style


def insert_zhr_svarki(def_db,road_programm,road_db,tube, km_start, km_finish, dy_tube):
    wb_zhr = load_workbook(road_programm+road_to_excel)
    ws_l1,ws_l2,ws_l3=wb_zhr['Tittle'],wb_zhr['Context'],wb_zhr['Style']
    current_row = 6
    numb_index = 1
    numb_stik = 0
    #-----------------------Title----------
    ws_l1['A3'] = def_db[0]['otv'][2]
    ws_l1['A5'] = 'Устранение дефектов на секциях %s, %s-%s км, Ду %s мм.'%(tube,km_start,km_finish,dy_tube)
    ws_l1['S13'] = km_start + ' км'
    ws_l1['AB13'] = km_finish + ' км'
    ws_l1['S18'] = '%s %s %s'%(def_db[0]['otv'][1], def_db[0]['otv'][2], def_db[0]['otv'][0])
    ws_l1['AL23'] = def_db[0]['date'][0].strftime('%d.%m.%Y')+' года'
    ws_l1['AL25'] = def_db[-1]['date'][0].strftime('%d.%m.%Y')+' года'
    # -----------------------Table----------
    for remont in def_db:
        type_repair = 'Шлифовка'
        for defect_num in remont['defect'].values():
            if 'П1' in defect_num['type']:
                type_repair = 'П1'
                field_1 = '09Г2С \nТУ 1469-022-04690510-02'
                field_2 = 'полумуфта\nполумуфта'
                field_3 = '№-\n№-'
                field_4 = 'СТ1,СТ2'
                numb_stik += 2
                break
            elif 'уфт' in defect_num['type']:
                type_repair = 'П2'
                field_1 = '09Г2С \nТУ 1469-001-01297858-01'
                field_2 = 'муфта\nсекция'
                field_3 = '№123 \n № ' + remont['sec']
                field_4 = 'СТ1,СТ2\nСТ3,СТ4\nСТ5,СТ6\nСТ7,СТ8\nСТ9,СТ10'
                numb_stik += 10
                break
        if type_repair != 'Шлифовка':
            create_row(ws_l2,ws_l3,current_row,(1,2),height=60)
            ws_l2['A' + str(current_row)] = numb_index
            ws_l2['B' + str(current_row)] = remont['date'][0].strftime('%d.%m.%y')+' г.'
            ws_l2['C' + str(current_row)] = '30 С°'
            ws_l2['D' + str(current_row)] = dy_tube
            ws_l2['E' + str(current_row)] = '8х8'
            ws_l2['F' + str(current_row)] = field_1
            ws_l2['G' + str(current_row)] = str(random.randint(105,120))+' С°'
            ws_l2['H' + str(current_row)] = field_2
            ws_l2['I' + str(current_row)] = field_3
            ws_l2['J' + str(current_row)] = str(remont['dl_muft']) + ' м\n№ - '
            ws_l2['K' + str(current_row)] = remont['km'] + ' км'
            ws_l2['L' + str(current_row)] = numb_index
            ws_l2['M' + str(current_row)] = field_4
            ws_l2['N' + str(current_row)] = remont['km'] + ' км'
            ws_l2['O' + str(current_row)] = '-'
            ws_l2['P' + str(current_row)] = 'РДС'
            ws_l2['Q' + str(current_row)] = ws_l2['AU1'].value
            ws_l2['R' + str(current_row)] = 'Красюков \n Схема №'
            ws_l2['S' + str(current_row)] = 'X\nY'
            #ws_l2['T' + str(current_row)] =
            ws_l2['U' + str(current_row)] = '№001-ВИК-{0}-0388-{1}-{2}'.format(type_repair,
                                                                               remont['km'],
                                                                               remont['date'][0].strftime('%d%m%y'))
            ws_l2['V' + str(current_row)] = remont['date'][0].strftime('%d.%m.%y') + ' г.'
            ws_l2['W' + str(current_row)] = 'годен'
            ws_l2['X' + str(current_row)] = 'УК'
            ws_l2['Y' + str(current_row)] = '№001-УК-{0}-0388-{1}-{2}'.format(type_repair,
                                                                               remont['km'],
                                                                               remont['date'][0].strftime('%d%m%y'))
            ws_l2['Z' + str(current_row)] = ws_l2['V' + str(current_row)].value
            ws_l2['AA' + str(current_row)] = 'годен'
            if type_repair == 'П2':
                ws_l2['AB' + str(current_row)] = 'ПВК'
                ws_l2['AC' + str(current_row)] = '№001-ПВК-{0}-0388-{1}-{2}'.format(type_repair,
                                                                                    remont['km'],
                                                                                    remont['date'][0].strftime('%d%m%y'))
                ws_l2['AD' + str(current_row)] = ws_l2['V' + str(current_row)].value
                ws_l2['AE' + str(current_row)] = 'годен'
            current_row += 1
            numb_index += 1
    #-----------------------Заключение

    ws_l1['AP46']='Работы закончены. Журнал закрыт. Сварено %s стыков. Отремонтированно 0(ноль) стыков.'%numb_stik
    ws_l1['AP47']='%s %s %s _____________ %s г.'%(remont['otv'][1],remont['otv'][2],remont['otv'][0],remont['date'][0].strftime('%d.%m.%Y'))
    ws_l2.print_area = 'A1:AT'+str(current_row)
    del(wb_zhr['Style'])
    road_db = os.path.normpath(road_db)
    road_db = os.path.join(road_db, 'Журналы')
    if not os.path.exists(road_db):
        os.makedirs(road_db)
    road_db = os.path.join(road_db, '4. Журнал сварки.xlsx')
    wb_zhr.save(road_db)

road_to_excel='/Excell/zhr_svarki.xlsx'
if __name__=='__main__':
    road_to_excel='../Excell/zhr_svarki.xlsx'
    road_db = os.getcwd() + '\..\..\Тест'
    road_programm = ''
    tube,km_start,km_finish,dy_tube='МНПП "Рязань-Тула-Орел" отвод на новомосковскую НБ, ДТ','12','13','530'
    def_db=[{'sec': '12',
             'date': [date(2018, 12, 12), date(2018, 12, 13), date(2018, 12, 14)],
             'dist': [10.1, 9.11, 11.09, 5.44, 14.76],
             'km': '1',
             'dl_muft': 1.0,
             'rand_value': [0.48533824224793676, 3.6729518931973657, 2.861346564462221, 2.705419511460805],
             'defect': {'1': {'sec': '12', 'dist': '123', 'lab': 'Риска с ППОШ с расслоением', 'dl': '5', 'sh': '6', 'gl': '7', 'type': 'Муфта П1'}, '2': {'sec': '12', 'dist': '123', 'lab': 'Риска с ППОШ с расслоением и ВНП', 'dl': '234', 'sh': '234', 'gl': '65', 'type': 'Муфта П2'}, '3': {'sec': '12', 'dist': '534536', 'lab': 'Риска с ППОШ с расслоением и ВНП', 'dl': '45534', 'sh': '345345', 'gl': '3535', 'type': 'Шлифовка'}}, 'otv': ('Афанасьев А.Б.', 'Начальник РУ№1', 'ЦРС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег"'), 'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
            {'sec': '13', 'date': [date(2018, 12, 13), date(2018, 12, 14), date(2018, 12, 14)], 'dist': [20.0, 18.54, 21.46, 15.55, 24.45], 'km': '2', 'dl_muft': 2.0, 'rand_value': [0.4581174447339836, 2.989229650534175, 2.9019640191285303, 2.8115670668469814], 'defect': {'1': {'sec': '13', 'dist': '12313', 'lab': 'Потеря метала', 'dl': '12', 'sh': '132', 'gl': '123', 'type': 'Шлифовка'}, '2': {'sec': '13', 'dist': '2344234', 'lab': 'Потеря метала', 'dl': '34', 'sh': '4343', 'gl': '434', 'type': 'Шлифовка'}}, 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('', '', ''), 'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
            {'sec': '14', 'date': [date(2018, 12, 14), date(2018, 12, 15), date(2018, 12, 16)], 'dist': [30.0, 27.94, 32.06, 27.74, 32.26], 'km': '3', 'dl_muft': 3.0, 'rand_value': [0.5619183603415756, 3.19813104261683, 3.0060601832822327, 2.9168215194726352], 'defect': {'1': {'sec': '14', 'dist': '1131213,32', 'lab': 'Риска', 'dl': '1233', 'sh': '12331', 'gl': '2133', 'type': 'Муфта П1'}, '2': {'sec': '14', 'dist': '321313,21', 'lab': 'Риска', 'dl': '23', 'sh': '234', 'gl': '2342', 'type': 'Муфта П1'}}, 'otv': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег'), 'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
            {'sec': '15', 'date': [date(2018, 12, 14), date(2018, 12, 16), date(2018, 12, 16)], 'dist': [160.0, 157.71, 162.29, 154.79, 165.21], 'km': '16', 'dl_muft': 4.0, 'rand_value': [0.2872474356253941, 2.915977705201697, 2.9123558666117617, 2.703436382412381], 'defect': {'1': {'sec': '15', 'dist': '123131,12', 'lab': 'Риска', 'dl': '123', 'sh': '123', 'gl': '13', 'type': 'Муфта П2'}, '2': {'sec': '15', 'dist': '1231', 'lab': 'Риска', 'dl': '3213', 'sh': '123', 'gl': '312', 'type': 'Муфта П2'}, '3': {'sec': '15', 'dist': '432423', 'lab': 'Риска', 'dl': '13', 'sh': '123', 'gl': '13', 'type': 'Муфта П2'}}, 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('', '', ''), 'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')}]
    insert_zhr_svarki(def_db, road_programm, road_db, tube, km_start, km_finish, dy_tube)
    
