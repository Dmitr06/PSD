from openpyxl import Workbook,load_workbook
from copy import copy
from datetime import *
import random,os

def create_row (ws,ws2,ws3,current_row,coord,height=13.50):
    current_row_2= current_row - 34
    ws.insert_rows(current_row)
    ws2.insert_rows(current_row_2)
    ws.row_dimensions[current_row].height = height
    ws2.row_dimensions[current_row_2].height = height
    ws.merge_cells('A{0}:F{0}'.format(current_row))
    ws.merge_cells('G{0}:M{0}'.format(current_row))
    ws.merge_cells('N{0}:T{0}'.format(current_row))
    ws.merge_cells('U{0}:Z{0}'.format(current_row))
    ws.merge_cells('AA{0}:AF{0}'.format(current_row))
    ws.merge_cells('AG{0}:AL{0}'.format(current_row))
    ws.merge_cells('AM{0}:AR{0}'.format(current_row))
    ws.merge_cells('AS{0}:AX{0}'.format(current_row))
    ws.merge_cells('AY{0}:BD{0}'.format(current_row))
    ws.merge_cells('BE{0}:BK{0}'.format(current_row))
    ws.merge_cells('BL{0}:CA{0}'.format(current_row))
    ws2.merge_cells('A{0}:F{0}'.format(current_row_2))
    ws2.merge_cells('G{0}:O{0}'.format(current_row_2))
    ws2.merge_cells('P{0}:X{0}'.format(current_row_2))
    ws2.merge_cells('Y{0}:AH{0}'.format(current_row_2))
    ws2.merge_cells('AI{0}:AQ{0}'.format(current_row_2))
    ws2.merge_cells('AR{0}:BD{0}'.format(current_row_2))
    ws2.merge_cells('BE{0}:BQ{0}'.format(current_row_2))
    ws2.merge_cells('BR{0}:CA{0}'.format(current_row_2))
    for index in range(1,80):
        ws.cell(current_row,index)._style = ws3.cell(*coord)._style
        ws2.cell(current_row_2,index)._style = ws3.cell(*coord)._style


def insert_zhr_ground(def_db,road_programm,road_db,tube,km_start,km_finish,dy_tube):
    wb_zhr = load_workbook(road_programm+road_to_excel)
    ws_l1,ws_l2,ws_l3=wb_zhr['L1'],wb_zhr['L2'],wb_zhr['Styles']
    current_row = 38
    current_row_2 = 4
    #-----------------------Title----------
    ws_l1['D4'] = def_db[0]['otv'][2]
    ws_l1['A15'] = 'Устранение дефектов на секциях %s, %s-%s км, Ду %s мм.'%(tube,km_start,km_finish,dy_tube)
    ws_l1['BO28'] = ws_l1['D4'].value
    ws_l1['BO30'] = def_db[0]['date'][0].strftime('%d.%m.%Y')+' года'
    ws_l1['BO32'] = def_db[-1]['date'][2].strftime('%d.%m.%Y')+' года'
    #-----------------------Table----------
    day_s=[]
    for remont in def_db:       #creat date list
        date_1=remont['date'][0]
        date_2=remont['date'][2]
        for date_num in (date_1,date_2):
            if date_num not in day_s:
                day_s.append(date_num)
    day_s=sorted(day_s)         #sort date list
    for i in day_s:
        for remont in def_db:
            V=round(remont['rand_value'][2]*remont['rand_value'][3]*(remont['dist'][4]-remont['dist'][3])*1.4)
            if remont['date'][0]==i:    #Excavation
                create_row(ws_l1,ws_l2,ws_l3,current_row,(1,1),height=62)
                ws_l1['A'+str(current_row)] = i.strftime('%d.%m.%Y')+' г.'  #date      
                ws_l1['G'+str(current_row)] = '%s км\nДист.\n%s м'%(remont['km'],remont['dist'][3]) #range start 
                ws_l1['N'+str(current_row)] = '%s км\nДист.\n%s м'%(remont['km'],remont['dist'][4]) #range stop
                ws_l1['U'+str(current_row)] = round(remont['dist'][4]-remont['dist'][3],1) #length
                ws_l1['AA'+str(current_row)] = round((ws_l1['U'+str(current_row)].value+20)*(remont['rand_value'][2]+20)) #planning works
                ws_l1['AG'+str(current_row)] =  round(V/13.3)
                ws_l1['AM'+str(current_row)] = V - ws_l1['AG'+str(current_row)].value
                ws_l1['AS'+str(current_row)] = '-'
                ws_l1['AY'+str(current_row)] = '-'
                ws_l1['BE'+str(current_row)] = '-'
                ws_l1['BL'+str(current_row)] = '-'
                ws_l2['A'+str(current_row_2)] = '%s/%s'%(round(remont['rand_value'][2],1),round(remont['rand_value'][3],1))
                ws_l2['G'+str(current_row_2)] = '-'
                ws_l2['P'+str(current_row_2)] = '-'
                ws_l2['AR'+str(current_row_2)] = '%s %s\n__________________\n%s'%(remont['otv'][1],remont['otv'][2],remont['otv'][0])
                ws_l2['BE'+str(current_row_2)] = '%s %s\n__________________\n%s'%(remont['contr'][1],remont['contr'][2],remont['contr'][0])
                current_row +=1
                current_row_2 +=1
            elif remont['date'][2]==i:  #covering
                create_row(ws_l1,ws_l2,ws_l3,current_row,(1,1),height=62)
                ws_l1['A'+str(current_row)]=i.strftime('%d.%m.%Y')+' г.'    #date
                ws_l1['G'+str(current_row)] = '%s км\nДист.\n%s м'%(remont['km'],remont['dist'][3]) #range start 
                ws_l1['N'+str(current_row)] = '%s км\nДист.\n%s м'%(remont['km'],remont['dist'][4]) #range stop
                ws_l1['U'+str(current_row)] = round(remont['dist'][4]-remont['dist'][3],1) #length
                ws_l1['AA'+str(current_row)] = '-'
                ws_l1['AG'+str(current_row)] = '-'
                ws_l1['AM'+str(current_row)] = '-'
                ws_l1['AS'+str(current_row)] = V
                ws_l1['AY'+str(current_row)] = round((ws_l1['U'+str(current_row)].value+20)*(remont['rand_value'][2]+20)) #planning works
                ws_l1['BE'+str(current_row)] = '-'
                ws_l1['BL'+str(current_row)] = '-'
                ws_l2['A'+str(current_row_2)] = '%s/%s'%(round(remont['rand_value'][2],1),round(remont['rand_value'][3],1))
                ws_l2['G'+str(current_row_2)] = 'СТТ'
                ws_l2['P'+str(current_row_2)] = 0.2
                ws_l2['AR'+str(current_row_2)] = '%s %s\n__________________\n%s'%(remont['otv'][1],remont['otv'][2],remont['otv'][0])
                ws_l2['BE'+str(current_row_2)] = '%s %s\n__________________\n%s'%(remont['contr'][1],remont['contr'][2],remont['contr'][0])
                current_row +=1
                current_row_2 +=1
    #-----------------------Заключение
    ws_l2['BY'+str(current_row_2+1)]='Работы закончены. Журнал закрыт.'
    ws_l2['BY'+str(current_row_2+2)]='%s %s %s _____________ %s г.'%(remont['otv'][1],remont['otv'][2],remont['otv'][0],remont['date'][2].strftime('%d.%m.%Y'))
    ws_l2.print_area = 'A1:CA'+str(current_row_2+3)
    ws_l1.print_area = 'A1:CA'+str(current_row)
    del(wb_zhr['Styles'])
    road=road_db.rpartition('/')[0]+'/Журналы'
    if not os.path.exists(road):
        os.makedirs(road)    
    wb_zhr.save(road+'/6. Журнал земляных работ.xlsx')

road_to_excel='/Excell/zhr_ground.xlsx'
if __name__=='__main__':
    import os
    road_to_excel='../Excell/zhr_ground.xlsx'
    road=os.getcwd()
    road_db=road.replace('\\','/')
    road_programm = ''
    tube,km_start,km_finish,dy_tube='МНПП "Рязань-Тула-Орел" отвод на новомосковскую НБ, ДТ','12','13','530'
    def_db=[{'sec': '12', 'date': [date(2018, 12, 12), date(2018, 12, 13), date(2018, 12, 14)], 'dist': [10.1, 9.11, 11.09, 5.44, 14.76], 'km': '1', 'dl_muft': 1.0, 'rand_value': [0.48533824224793676, 3.6729518931973657, 2.861346564462221, 2.705419511460805], 'defect': {'1': {'sec': '12', 'dist': '123', 'lab': 'Риска с ППОШ с расслоением', 'dl': '5', 'sh': '6', 'gl': '7', 'type': 'Муфта П1'}, '2': {'sec': '12', 'dist': '123', 'lab': 'Риска с ППОШ с расслоением и ВНП', 'dl': '234', 'sh': '234', 'gl': '65', 'type': 'Муфта П2'}, '3': {'sec': '12', 'dist': '534536', 'lab': 'Риска с ППОШ с расслоением и ВНП', 'dl': '45534', 'sh': '345345', 'gl': '3535', 'type': 'Шлифовка'}}, 'otv': ('Афанасьев А.Б.', 'Начальник РУ№1', 'ЦРС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег"'), 'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
            {'sec': '13', 'date': [date(2018, 12, 13), date(2018, 12, 14), date(2018, 12, 14)], 'dist': [20.0, 18.54, 21.46, 15.55, 24.45], 'km': '2', 'dl_muft': 2.0, 'rand_value': [0.4581174447339836, 2.989229650534175, 2.9019640191285303, 2.8115670668469814], 'defect': {'1': {'sec': '13', 'dist': '12313', 'lab': 'Потеря метала', 'dl': '12', 'sh': '132', 'gl': '123', 'type': 'Шлифовка'}, '2': {'sec': '13', 'dist': '2344234', 'lab': 'Потеря метала', 'dl': '34', 'sh': '4343', 'gl': '434', 'type': 'Шлифовка'}}, 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('', '', ''), 'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
            {'sec': '14', 'date': [date(2018, 12, 14), date(2018, 12, 15), date(2018, 12, 16)], 'dist': [30.0, 27.94, 32.06, 27.74, 32.26], 'km': '3', 'dl_muft': 3.0, 'rand_value': [0.5619183603415756, 3.19813104261683, 3.0060601832822327, 2.9168215194726352], 'defect': {'1': {'sec': '14', 'dist': '1131213,32', 'lab': 'Риска', 'dl': '1233', 'sh': '12331', 'gl': '2133', 'type': 'Муфта П1'}, '2': {'sec': '14', 'dist': '321313,21', 'lab': 'Риска', 'dl': '23', 'sh': '234', 'gl': '2342', 'type': 'Муфта П1'}}, 'otv': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег'), 'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
            {'sec': '15', 'date': [date(2018, 12, 14), date(2018, 12, 16), date(2018, 12, 16)], 'dist': [160.0, 157.71, 162.29, 154.79, 165.21], 'km': '16', 'dl_muft': 4.0, 'rand_value': [0.2872474356253941, 2.915977705201697, 2.9123558666117617, 2.703436382412381], 'defect': {'1': {'sec': '15', 'dist': '123131,12', 'lab': 'Риска', 'dl': '123', 'sh': '123', 'gl': '13', 'type': 'Муфта П2'}, '2': {'sec': '15', 'dist': '1231', 'lab': 'Риска', 'dl': '3213', 'sh': '123', 'gl': '312', 'type': 'Муфта П2'}, '3': {'sec': '15', 'dist': '432423', 'lab': 'Риска', 'dl': '13', 'sh': '123', 'gl': '13', 'type': 'Муфта П2'}}, 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('', '', ''), 'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')}]
    insert_zhr_ground(def_db,road_programm,road_db,tube,km_start,km_finish,dy_tube)
    
