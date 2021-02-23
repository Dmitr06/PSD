from openpyxl import load_workbook
from datetime import *

def create_row (ws,ws2,ws3,current_row,coord,height=13.50):
    current_row_2= current_row - 34
    ws.insert_rows(current_row)
    ws2.insert_rows(current_row_2)
    ws.row_dimensions[current_row].height = height
    ws2.row_dimensions[current_row_2].height = height
    ws.merge_cells('A{0}:E{0}'.format(current_row))
    ws.merge_cells('F{0}:N{0}'.format(current_row))
    ws.merge_cells('O{0}:Q{0}'.format(current_row))
    ws.merge_cells('R{0}:V{0}'.format(current_row))
    ws.merge_cells('W{0}:AK{0}'.format(current_row))
    ws.merge_cells('AL{0}:AO{0}'.format(current_row))
    ws.merge_cells('AP{0}:AT{0}'.format(current_row))
    ws.merge_cells('AU{0}:BE{0}'.format(current_row))
    ws.merge_cells('BF{0}:BZ{0}'.format(current_row))
    ws2.merge_cells('A{0}:E{0}'.format(current_row_2))
    ws2.merge_cells('F{0}:P{0}'.format(current_row_2))
    ws2.merge_cells('Q{0}:T{0}'.format(current_row_2))
    ws2.merge_cells('U{0}:AA{0}'.format(current_row_2))
    ws2.merge_cells('AB{0}:AR{0}'.format(current_row_2))
    ws2.merge_cells('AS{0}:BC{0}'.format(current_row_2))
    ws2.merge_cells('BD{0}:BN{0}'.format(current_row_2))
    ws2.merge_cells('BO{0}:BZ{0}'.format(current_row_2))
    for index in range(1,79):
        ws.cell(current_row,index)._style = ws3.cell(*coord)._style
        ws2.cell(current_row_2,index)._style = ws3.cell(*coord)._style


def insert_zhr_izol(def_db,road_programm,road_db,tube,km_start,km_finish,dy_tube):
    wb_zhr = load_workbook(road_programm+road_to_excel)
    ws_l1,ws_l2,ws_l3=wb_zhr['L1'],wb_zhr['L2'],wb_zhr['Styles']
    current_row = 39
    current_row_2 = 5
    #-----------------------Title----------
    ws_l1['D4'] = def_db[0]['otv'][2]
    ws_l1['A15'] = 'Устранение дефектов на секциях %s, %s-%s км, Ду %s мм.'%(tube,km_start,km_finish,dy_tube)
    ws_l1['BN28'] = ws_l1['D4'].value
    ws_l1['BN30'] = def_db[0]['date'][0].strftime('%d.%m.%Y')+' года'
    ws_l1['BN32'] = def_db[-1]['date'][2].strftime('%d.%m.%Y')+' года'
    #-----------------------Table----------
    for remont in def_db:
        create_row(ws_l1,ws_l2,ws_l3,current_row,(1,1),height=81)
        ws_l1['A'+str(current_row)] = remont['date'][1].strftime('%d.%m.%y')+' г.'
        ws_l1['F'+str(current_row)] = '%s, %s км.'%(tube,remont['km'])
        ws_l1['O'+str(current_row)] = ' С°'
        ws_l1['R'+str(current_row)] = '2'
        ws_l1['W'+str(current_row)] = ws_l3['A3'].value
        ws_l1['AL'+str(current_row)] = '40 С°'
        ws_l1['AP'+str(current_row)] = '-'
        ws_l1['AU'+str(current_row)] = 'Соотвествует ГОСТ Р 51164-98'
        ws_l1['BF'+str(current_row)] = 'Укладка не проводилась'
        ws_l2['F'+str(current_row_2)] = 'Ремонт после проверки адгезии\n\n_______%s\n%s'%('Лазарев А.В.',remont['date'][2].strftime('%d.%m.%y')+' г.')
        ws_l2['Q'+str(current_row_2)] = ' С°'
        ws_l2['U'+str(current_row_2)] = remont['date'][2].strftime('%d.%m.%Y')+' г.'
        ws_l2['AS'+str(current_row_2)] = 'Мастер РУ №2 ЦРС "Рязань"\n\n_______________\nЛазарев А.В.'
        ws_l2['BD'+str(current_row_2)] = '%s %s\n\n__________________\n%s'%(remont['otv'][1],remont['otv'][2],remont['otv'][0])
        ws_l2['BO'+str(current_row_2)] = '%s %s\n\n__________________\n%s'%(remont['contr'][1],remont['contr'][2],remont['contr'][0])
        current_row += 1
        current_row_2 +=1
  
    #-----------------------Заключение
    ws_l2['BX'+str(current_row_2+1)]='Работы закончены. Журнал закрыт.'
    ws_l2['BX'+str(current_row_2+2)]='%s %s %s _____________ %s г.'%(remont['otv'][1],remont['otv'][2],remont['otv'][0],remont['date'][2].strftime('%d.%m.%Y'))
    ws_l2.print_area = 'A1:BZ'+str(current_row_2+3)
    ws_l1.print_area = 'A1:BZ'+str(current_row)
    del(wb_zhr['Styles'])
    road=road_db.rpartition('/')[0]+'/Журналы'
    if not os.path.exists(road):
        os.makedirs(road)    
    wb_zhr.save(road+'/5. Журнал изоляции.xlsx')

road_to_excel='/Excell/zhr_izol.xlsx'
if __name__=='__main__':
    import os
    road_to_excel='/../Excell/zhr_izol.xlsx'
    road=os.getcwd()
    road=road.replace('\\','/')
    tube,km_start,km_finish,dy_tube='МНПП "Рязань-Тула-Орел" отвод на новомосковскую НБ, ДТ','12','13','530'
    def_db=[{'sec': '12', 'date': [date(2018, 12, 12), date(2018, 12, 13), date(2018, 12, 14)], 'dist': [10.1, 9.11, 11.09, 5.44, 14.76], 'km': '1', 'dl_muft': 1.0, 'rand_value': [0.48533824224793676, 3.6729518931973657, 2.861346564462221, 2.705419511460805], 'defect': {'1': {'sec': '12', 'dist': '123', 'lab': 'Риска с ППОШ с расслоением', 'dl': '5', 'sh': '6', 'gl': '7', 'type': 'Муфта П1'}, '2': {'sec': '12', 'dist': '123', 'lab': 'Риска с ППОШ с расслоением и ВНП', 'dl': '234', 'sh': '234', 'gl': '65', 'type': 'Муфта П2'}, '3': {'sec': '12', 'dist': '534536', 'lab': 'Риска с ППОШ с расслоением и ВНП', 'dl': '45534', 'sh': '345345', 'gl': '3535', 'type': 'Шлифовка'}}, 'otv': ('Афанасьев А.Б.', 'Начальник РУ№1', 'ЦРС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег"'), 'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
            {'sec': '13', 'date': [date(2018, 12, 13), date(2018, 12, 14), date(2018, 12, 14)], 'dist': [20.0, 18.54, 21.46, 15.55, 24.45], 'km': '2', 'dl_muft': 2.0, 'rand_value': [0.4581174447339836, 2.989229650534175, 2.9019640191285303, 2.8115670668469814], 'defect': {'1': {'sec': '13', 'dist': '12313', 'lab': 'Потеря метала', 'dl': '12', 'sh': '132', 'gl': '123', 'type': 'Шлифовка'}, '2': {'sec': '13', 'dist': '2344234', 'lab': 'Потеря метала', 'dl': '34', 'sh': '4343', 'gl': '434', 'type': 'Шлифовка'}}, 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('', '', ''), 'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
            {'sec': '14', 'date': [date(2018, 12, 14), date(2018, 12, 15), date(2018, 12, 16)], 'dist': [30.0, 27.94, 32.06, 27.74, 32.26], 'km': '3', 'dl_muft': 3.0, 'rand_value': [0.5619183603415756, 3.19813104261683, 3.0060601832822327, 2.9168215194726352], 'defect': {'1': {'sec': '14', 'dist': '1131213,32', 'lab': 'Риска', 'dl': '1233', 'sh': '12331', 'gl': '2133', 'type': 'Муфта П1'}, '2': {'sec': '14', 'dist': '321313,21', 'lab': 'Риска', 'dl': '23', 'sh': '234', 'gl': '2342', 'type': 'Муфта П1'}}, 'otv': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег'), 'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
            {'sec': '15', 'date': [date(2018, 12, 14), date(2018, 12, 16), date(2018, 12, 16)], 'dist': [160.0, 157.71, 162.29, 154.79, 165.21], 'km': '16', 'dl_muft': 4.0, 'rand_value': [0.2872474356253941, 2.915977705201697, 2.9123558666117617, 2.703436382412381], 'defect': {'1': {'sec': '15', 'dist': '123131,12', 'lab': 'Риска', 'dl': '123', 'sh': '123', 'gl': '13', 'type': 'Муфта П2'}, '2': {'sec': '15', 'dist': '1231', 'lab': 'Риска', 'dl': '3213', 'sh': '123', 'gl': '312', 'type': 'Муфта П2'}, '3': {'sec': '15', 'dist': '432423', 'lab': 'Риска', 'dl': '13', 'sh': '123', 'gl': '13', 'type': 'Муфта П2'}}, 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('', '', ''), 'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')}]
    insert_zhr_izol(def_db,road,road,tube,km_start,km_finish,dy_tube)
    
