from openpyxl import Workbook,load_workbook
from openpyxl.styles import Border,Side,Font
from copy import copy
import os


def insert_zhr_vh(akt_vh,road_programm,road_db,tube,km_start,km_finish,dy_tube):
    wb_zhr = load_workbook(road_programm+road_to_excel)
    ws_l1,ws_l2=wb_zhr['L1'],wb_zhr['L2']

    #-----------------------Титульный лист журнала
    ws_l1['D3'] = akt_vh[0]['otv'][2]
    ws_l1['AC3'] = 'Устранение дефектов на секциях %s, %s-%s км, Ду %s мм.'%(tube,km_start,km_finish,dy_tube)
    ws_l1['AP7'] = km_start+' км.'
    ws_l1['AP9'] = km_finish+' км.'
    ws_l1['A16'] = ws_l1['AC3'].value
    ws_l1['AL28'] = akt_vh[0]['date']+' г.'
    ws_l1['AL30'] = akt_vh[-1]['date']+' г.'
    #-----------------------Материалы
    num_start=41 # начинаем заполнять с данной строки
    col_stand=['P','R','X','AL']
    col_change=['A','C','I','N','AF','AI']
    for i in range(39,41):
        ws_l1['A'+str(i)]=akt_vh[0]['date']+' г.'
        ws_l1['AF'+str(i)]='%s %s\n_________________\n%s'%(akt_vh[0]['otv'][1],akt_vh[0]['otv'][2],akt_vh[0]['otv'][0])
        ws_l1['AI'+str(i)]='%s %s\n_________________\n%s'%(akt_vh[0]['contr'][1],akt_vh[0]['contr'][2],akt_vh[0]['contr'][0])
    #------Создаем поля согласно бд
    for x in akt_vh:
        ws_l1.merge_cells('A{0}:B{0}'.format(num_start))
        ws_l1.merge_cells('C{0}:H{0}'.format(num_start))
        ws_l1.merge_cells('I{0}:M{0}'.format(num_start))
        ws_l1.merge_cells('N{0}:O{0}'.format(num_start))
        ws_l1.merge_cells('P{0}:Q{0}'.format(num_start))
        ws_l1.merge_cells('R{0}:W{0}'.format(num_start))
        ws_l1.merge_cells('X{0}:AE{0}'.format(num_start))
        ws_l1.merge_cells('AF{0}:AH{0}'.format(num_start)) 
        ws_l1.merge_cells('AI{0}:AK{0}'.format(num_start))
        ws_l1.merge_cells('AL{0}:AS{0}'.format(num_start))
        for y in range(1,46):
            ws_l1.cell(num_start,y)._style = ws_l1.cell(40,y)._style
        index_mat=1
        ws_l1.row_dimensions[num_start].height = 100
        #------Вставляем занчение из бд в поля
        temp=str(num_start)
        if any(j in x['full_name'] for j in ('Литкор','Полилен')):
            index_mat=2
            ws_l1.row_dimensions[num_start].height = 130
        elif any(j in x['full_name'] for j in ('герм','композ','Герм','Композ')):
            index_mat=3
        elif any(j in x['full_name'] for j in ('муфт','Муфт')):
            index_mat=4
        for i in col_stand:
            ws_l1[i+temp]=ws_l2[i+str(index_mat)].value
        ws_l1['AL'+temp].value=ws_l1['AL'+temp].value.replace('MATERIAL',x['full_name'])
        ws_l1['A'+temp]=x['date']+' г.' # 
        ws_l1['C'+temp]=x['full_name'] #
        ws_l1['I'+temp]=x['doc'] #
        ws_l1['N'+temp]='%s %s.'%(x['kol'],x['marker']) #
        ws_l1['AF'+temp]='%s %s\n_________________\n%s'%(x['otv'][1],x['otv'][2],x['otv'][0]) # 
        ws_l1['AI'+temp]='%s %s\n_________________\n%s'%(x['contr'][1],x['contr'][2],x['contr'][0]) # 
        num_start+=1

    #-----------------------Заключение
    ft = Font(name='Times New Roman')
    ws_l1['AF'+str(num_start+2)].font = ft
    ws_l1['AF'+str(num_start+2)]='Работы закончены. Журнал закрыт'
    temp=akt_vh[-1]
    ws_l1['U'+str(num_start+3)].font = ft
    ws_l1['U'+str(num_start+3)]='%s %s %s _____________ %s г.'%(temp['otv'][1],temp['otv'][2],temp['otv'][0],temp['date'])
    ws_l1.print_area = 'A1:AS'+str(num_start+4)
    wb_zhr.remove_sheet(wb_zhr['L2'])
    wb_zhr.remove_sheet(wb_zhr['L3'])
    road=road_db.rpartition('/')[0]+'/Входной контроль'
    if not os.path.exists(road):
        os.makedirs(road)    
    wb_zhr.save(road+'/Журнал входного контроля.xlsx')

road_to_excel='/Excell/zhr_vh.xlsx'
if __name__=='__main__':
    import os
    road_to_excel='../Excell/zhr_vh.xlsx'
    road=os.getcwd()
    road_db=road.replace('\\','/')
    road_programm = ''
    tube,km_start,km_finish,dy_tube='МНПП "Рязань-Тула-Орел" отвод на новомосковскую НБ, ДТ','12','13','530'
    akt_vh=[{'name': 'Абразив', 'full_name': 'Абразив 123', 'marker': 'кг', 'TU': 'ТУ5', 'param': '-', 'date': '02.03.2019', 'doc': 'Паспорт 34234', 'kol': '3', 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'), 'contr': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'), 'kk': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег')},
            {'name': 'Полилен-ПЛ', 'full_name': ' материал Полилен-ПЛ', 'marker': 'рул.', 'TU': 'ТУ3', 'param': 'длина= мм, ширина= мм', 'date': '01.12.2019', 'doc': 'Паспорт 14', 'kol': '1', 'otv': ('Афанасьев А.Б.', 'Начальник РУ№1', 'ЦРС "Рязань"'), 'contr': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'), 'kk': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег')},
            {'name': 'Муфта П1 530', 'full_name': 'Муфта КМТ 530', 'marker': 'шт.', 'TU': 'ТУ8', 'param': 'диаметр= мм, длина= мм, толщина стенки= м', 'date': '05.12.2019', 'doc': 'Паспорт 12', 'kol': '3534', 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'kk': ('', '', ''), 'sk': ('', '', '')},
            {'name': 'Герметик', 'full_name': 'состав полмерный Герметик', 'marker': 'компл.', 'TU': 'ТУ6', 'param': '-', 'date': '07.12.2019', 'doc': 'Паспорт 645', 'kol': '3', 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'kk': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег')},
            {'name': 'Полилен-ПЛ', 'full_name': ' материал Полилен-ПЛ', 'marker': 'рул.', 'TU': 'ТУ3', 'param': 'длина= мм, ширина= мм', 'date': '01.12.2019', 'doc': 'Паспорт 14', 'kol': '1', 'otv': ('Афанасьев А.Б.', 'Начальник РУ№1', 'ЦРС "Рязань"'), 'contr': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'), 'kk': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег')},
            {'name': 'Муфта П1 530', 'full_name': 'Муфта КМТ 530', 'marker': 'шт.', 'TU': 'ТУ8', 'param': 'диаметр= мм, длина= мм, толщина стенки= м', 'date': '05.12.2019', 'doc': 'Паспорт 12', 'kol': '3534', 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'kk': ('', '', ''), 'sk': ('', '', '')},
            {'name': 'Герметик', 'full_name': 'состав полмерный Герметик', 'marker': 'компл.', 'TU': 'ТУ6', 'param': '-', 'date': '07.12.2019', 'doc': 'Паспорт -', 'kol': '3', 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'), 'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'kk': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег')}]   
    insert_zhr_vh(akt_vh,road_programm,road_db,tube,km_start,km_finish,dy_tube)
    
