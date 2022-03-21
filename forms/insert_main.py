from openpyxl import load_workbook
import re
from datetime import *


def create_row(ws, ws2, current_row, coord, height=13.50):
    ws.insert_rows(current_row)
    ws.row_dimensions[current_row].height = height
    ws.merge_cells('A{0}:N{0}'.format(current_row))
    ws.merge_cells('O{0}:AY{0}'.format(current_row))
    ws.merge_cells('AZ{0}:BG{0}'.format(current_row))
    ws.merge_cells('BH{0}:BQ{0}'.format(current_row))
    ws.merge_cells('BR{0}:BX{0}'.format(current_row))
    ws.merge_cells('BY{0}:CA{0}'.format(current_row))
    for index in range(1, 80):
        ws.cell(current_row, index)._style = ws2.cell(*coord)._style


def insert_main(def_db, road_programm, road_db, tube, km_start, km_finish, dy_tube):
    wb = load_workbook(road_programm + road_to_excel)
    ws_style = wb['Styles']
    # --------------Реестр
    ws1 = wb['Reestr']
    ws1['A1'] = def_db[0]['otv'][2]
    ws1['A3'] = 'Устранение дефектов методом выборочного ремонта  на секциях {0}, {1}-{2} км, Ду {3} мм. '.format(
        tube, km_start, km_finish, dy_tube)
    height_row = 20
    for i in def_db:
        height_row = height_row + 10
        ws1['CO2'] = ws1['CO2'].value + '\n' + i['zakl']
        ws1['CO3'] = ws1['CO3'].value + '\n' + i['zakl'].replace('В', 'У')
        ws1['CO4'] = ws1['CO4'].value + '\n' + i['zakl'].replace('В', 'К')
    '''ws1.row_dimensions[38].height = height_row
    ws1.row_dimensions[39].height = height_row
    ws1.row_dimensions[40].height = height_row'''
    ws1['CO5'] = '№' + def_db[0]['sec']
    for i in def_db[1:]:
        ws1['CO5'] = ws1['CO5'].value + ', ' + i['sec']
        if i['otv'][2] not in ws1['A1'].value:
            ws1['A1'] = ws1['A1'].value + ', ' + i['otv'][2]
    ws1['CO6'] = '№1'
    if len(def_db) > 1:
        ws1['CO6'] = ws1['CO6'].value + '-%s' % (len(def_db))
    # --------------Перечень лиц
    ws2 = wb['Perechen']
    index_row = 14
    contr = []
    lkk = []
    sk = []
    otv = []
    for defect in def_db:
        for (x, y, z) in zip((defect['contr'], defect['lkk'], defect['sk'], defect['otv']), (contr, lkk, sk, otv),
                             ('A2', 'A3', 'A4', 'A5')):
            if (x not in y) and (x[0] != ''):
                y.append(x)
                if (y is contr) or (y is otv):
                    height_row = 150
                else:
                    height_row = 60
                create_row(ws2, ws_style, index_row, (1, 1), height_row)
                temp_row = str(index_row)
                ws2['A' + temp_row] = x[2]
                ws2['O' + temp_row] = ws_style[z].value
                if not (y is sk): ws2['AZ' + temp_row] = '=AZ13'
                ws2['BH' + temp_row] = '%s %s %s' % (x[1], x[2], x[0])
                index_row += 1
    index_row += 1
    temp_row = str(index_row)
    ws2['BO' + temp_row] = def_db[0]['date'][0].strftime('%d.%m.%Y') + ' г.'
    index_row += 1
    ws2.print_area = 'A1:CA' + str(index_row)
    # --------------Справка об очистке участка
    ws3 = wb['Musor']
    ws3['AY6'] = def_db[-1]['date'][2].strftime('%d.%m.%Y') + ' года'
    ws3['A13'] = re.sub(r'replace', '%s %s км-%s км' % (tube, km_start, km_finish), ws3['A13'].value)
    ws3['P39'] = '%s %s %s' % (def_db[-1]['contr'][1], def_db[-1]['contr'][2], def_db[-1]['contr'][0])
    ws3['AT39'] = def_db[-1]['date'][2].strftime('%d.%m.%Y') + ' г.'
    ws3['P42'] = '%s %s %s' % (def_db[-1]['otv'][1], def_db[-1]['otv'][2], def_db[-1]['otv'][0])
    ws3['AT42'] = def_db[-1]['date'][2].strftime('%d.%m.%Y') + ' г.'
    # --------------Справка об отсутствии замечаний
    ws4 = wb['Spravka']
    index_row = 12
    for section in def_db:
        ws4.insert_rows(index_row)
        temp_str = ', '.join(section['defect'].keys())
        ws4.cell(index_row, 49)._style = ws_style.cell(1, 2)._style
        ws4['AW' + str(index_row)] = '№%s дефектов №№%s;' % (section['sec'], temp_str)
        ws4.merge_cells('A{0}:AV{0}'.format(index_row))
        ws4.cell(index_row, 1)._style = ws_style.cell(1, 2)._style
        ws4['A' + str(index_row)] = '№%s дефектов №№%s;' % (section['sec'], temp_str)
        index_row += 1
    ws4.merge_cells('A{0}:AV{0}'.format(index_row))
    ws4.cell(index_row, 1)._style = ws_style.cell(1, 2)._style
    ws4['A' + str(index_row)] = "%s, %s км-%s км." % (tube, km_start, km_finish)
    index_row += 13
    ws4['U' + str(index_row)] = ' '.join((def_db[-1]['sk'][1], def_db[-1]['sk'][2], def_db[-1]['sk'][0]))
    ws4.print_area = 'A1:AV' + str(index_row + 8)
    # --------------Акт передачи
    ws5 = wb['Akt']
    index_row = 16
    for section in def_db:
        ws5.insert_rows(index_row)
        temp_str = ', '.join(section['defect'].keys())
        ws5.cell(index_row, 49)._style = ws_style.cell(1, 2)._style
        ws5['AW' + str(index_row)] = '№%s дефектов №№%s;' % (section['sec'], temp_str)
        ws5.merge_cells('A{0}:AV{0}'.format(index_row))
        ws5.cell(index_row, 1)._style = ws_style.cell(1, 2)._style
        ws5['A' + str(index_row)] = '№%s дефектов №№%s;' % (section['sec'], temp_str)
        index_row += 1
    ws5.merge_cells('A{0}:AV{0}'.format(index_row))
    ws5.cell(index_row, 1)._style = ws_style.cell(1, 2)._style
    ws5['A' + str(index_row)] = "%s, %s км-%s км." % (tube, km_start, km_finish)
    ws5.print_area = 'A1:AV' + str(index_row + 13)
    # ----------------Конец
    wb.remove(wb['Styles'])
    wb.save(road_db.rpartition('/')[0] + '/0. Основное.xlsx')


road_to_excel = '/Excell/main.xlsx'
if __name__ == '__main__':
    import os

    road_to_excel = '../Excell/main.xlsx'
    road = os.getcwd()
    road_db = road.replace('\\', '/')[:-9] + 'Тест/'
    print(road_db)
    road_programm = ''
    tube, km_start, km_finish, dy_tube = 'МНПП "Рязань-Тула-Орел" отвод на новомосковскую НБ, ДТ', '12', '13', '530'
    def_db = [{'sec': '12',
               'date': [date(2018, 12, 12), date(2018, 12, 13), date(2018, 12, 14)],
               'dist': [10.1, 9.11, 11.09, 5.44, 14.76],
               'km': '1',
               'zakl': '21/В-1',
               'dl_muft': 1.0,
               'rand_value': [0.48533824224793676, 3.6729518931973657, 2.861346564462221, 2.705419511460805],
               'defect': {'1': {'sec': '12',
                                'dist': '123',
                                'lab': 'Риска с ППОШ с расслоением',
                                'dl': '5',
                                'sh': '6',
                                'gl': '7',
                                'type': 'Муфта П1'},
                          '2': {'sec': '12',
                                'dist': '123',
                                'lab': 'Риска с ППОШ с расслоением и ВНП',
                                'dl': '234',
                                'sh': '234',
                                'gl': '65',
                                'type': 'Муфта П2'},
                          '3': {'sec': '12',
                                'dist': '534536',
                                'lab': 'Риска с ППОШ с расслоением и ВНП',
                                'dl': '45534',
                                'sh': '345345',
                                'gl': '3535',
                                'type': 'Шлифовка'}},
               'otv': ('Афанасьев А.Б.', 'Начальник РУ№1', 'ЦРС "Рязань"'),
               'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'),
               'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег'),
               'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
              {'sec': '13',
               'date': [date(2018, 12, 13),
                        date(2018, 12, 13),
                        date(2018, 12, 14)],
               'dist': [20.0, 18.54, 21.46, 15.55, 24.45],
               'km': '2',
               'zakl': '21/В-2',
               'dl_muft': 2.0,
               'rand_value': [0.4581174447339836, 2.989229650534175, 2.9019640191285303, 2.8115670668469814],
               'defect': {
                   '1': {'sec': '13', 'dist': '12313', 'lab': 'Потеря метала', 'dl': '12', 'sh': '132', 'gl': '123',
                         'type': 'Шлифовка'},
                   '2': {'sec': '13', 'dist': '2344234', 'lab': 'Потеря метала', 'dl': '34', 'sh': '4343', 'gl': '434',
                         'type': 'Шлифовка'}}, 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'),
               'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('', '', ''),
               'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
              {'sec': '14',
               'date': [date(2018, 12, 14),
                        date(2018, 12, 15),
                        date(2018, 12, 16)],
               'dist': [30.0, 27.94, 32.06, 27.74, 32.26],
               'km': '3',
               'zakl': '21/В-3',
               'dl_muft': 3.0,
               'rand_value': [0.5619183603415756, 3.19813104261683, 3.0060601832822327, 2.9168215194726352],
               'defect': {
                   '1': {'sec': '14', 'dist': '1131213,32', 'lab': 'Риска', 'dl': '1233', 'sh': '12331', 'gl': '2133',
                         'type': 'Муфта П1'},
                   '2': {'sec': '14', 'dist': '321313,21', 'lab': 'Риска', 'dl': '23', 'sh': '234', 'gl': '2342',
                         'type': 'Муфта П1'}}, 'otv': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'),
               'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'),
               'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег'),
               'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')},
              {'sec': '15',
               'date': [date(2018, 12, 14),
                        date(2018, 12, 14),
                        date(2018, 12, 15)],
               'dist': [160.0, 157.71, 162.29, 154.79, 165.21],
               'km': '16',
               'zakl': '21/В-4',
               'dl_muft': 4.0,
               'rand_value': [0.2872474356253941, 2.915977705201697, 2.9123558666117617, 2.703436382412381],
               'defect': {'1': {'sec': '15', 'dist': '123131,12', 'lab': 'Риска', 'dl': '123', 'sh': '123', 'gl': '13',
                                'type': 'Муфта П2'},
                          '2': {'sec': '15', 'dist': '1231', 'lab': 'Риска', 'dl': '3213', 'sh': '123', 'gl': '312',
                                'type': 'Муфта П2'},
                          '3': {'sec': '15', 'dist': '432423', 'lab': 'Риска', 'dl': '13', 'sh': '123', 'gl': '13',
                                'type': 'Муфта П2'}}, 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'),
               'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'sk': ('sdfsdf', '3213123', 'fsd'),
               'lkk': ('Козин А.П.', 'Инженер-дефектоскопист', 'ЛККиД')}]
    insert_main(def_db, road_programm, road_db, tube, km_start, km_finish, dy_tube)
