from openpyxl import load_workbook
from openpyxl.worksheet.copier import WorksheetCopy
import os


def insert_akt_vh(akt_vh, road_programm, road_db, tube, km_start, km_finish, dy_tube):
    wb = load_workbook(road_programm + road_to_excel)
    for i, x in enumerate(akt_vh):
        current_wb = wb.create_sheet('%s. %s' % (i + 1, x['name']))
        copy_ws = WorksheetCopy(wb['Empty'], current_wb)
        copy_ws.copy_worksheet()
        copy_ws._copy_cells()
        copy_ws._copy_dimensions()
        current_wb.print_area = 'A1:AB121'
        current_wb.page_margins = wb['Empty'].page_margins
        # --------------начинаем заполнение текущего акта
        current_wb['A1'] = x['otv'][2]  # служба
        current_wb['A3'] = 'Устранение дефектов на секциях %s, %s-%s км, Ду %s мм.' % (
        tube, km_start, km_finish, dy_tube)  # заголовок
        current_wb['K27'] = '%s, %s-%s км.' % (tube, km_start, km_finish)  # заголовок
        current_wb['M11'] = x['date'] + ' г.'  # дата
        current_wb['P6'] = i + 1  # номер акта
        current_wb['A8'] = '%s, в кол-ве %s %s.' % (x['full_name'], x['kol'], x['marker'])  # номер акта
        current_wb['K13'] = ' '.join((x['otv'][1], x['otv'][2], x['otv'][0]))
        current_wb['K19'] = ' '.join((x['contr'][1], x['contr'][2], x['contr'][0]))
        current_wb['K17'] = ' '.join((x['sk'][1], x['sk'][2], x['sk'][0]))
        current_wb['J32'] = x['doc']
        current_wb['A35'] = x['param']
        current_wb['H37'] = x['TU']
        if 'OVP' in x:
            current_wb['N47'] = x['OVP']
        else:
                        current_wb['A46'] = current_wb['A46'].value.replace('находится','не находится')
        current_wb['L51'] = x['otv'][0]
        current_wb['L57'] = x['contr'][0]
        current_wb['L55'] = x['sk'][0]
        current_wb.print_area = 'A1:AB58'
    del (wb['Empty'])
    road = road_db.rpartition('/')[0] + '/Входной контроль'
    if not os.path.exists(road):
        os.makedirs(road)
    wb.save(road + '/Акты входного контроля.xlsx')


road_to_excel = '/Excell/akt_vh.xlsx'
if __name__ == '__main__':
    road_to_excel = '../Excell/akt_vh.xlsx'
    road = os.getcwd()
    road_db=road.replace('\\','/')[:-9]+'Тест/'
    print(road_db)
    road_programm = ''
    tube, km_start, km_finish, dy_tube = 'МНПП "Рязань-Тула-Орел" отвод на новомосковскую НБ, ДТ', '12', '13', '530'
    akt_vh = [
        {'name': 'Абразив', 'full_name': 'Абразим 123', 'marker': 'кг', 'TU': 'ТУ5', 'param': '-', 'date': '02.03.2019',
         'doc': 'Паспорт 34234', 'kol': '3', 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'),
         'contr': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'),
         'kk': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег"')},
        {'name': 'Праймер', 'full_name': 'Праймер ПЛ', 'marker': 'рул', 'TU': 'ТУ3', 'param': 'длина= мм, ширина= мм',
         'date': '01.12.2019', 'doc': 'Сертификат 14', 'kol': '1',
         'otv': ('Афанасьев А.Б.', 'Начальник РУ№1', 'ЦРС "Рязань"'),
         'contr': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'),
         'kk': ('Макаров М.А.', 'Начальник АРС', 'ЛПДС "Рязань"'), 'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег"')},
        {'name': 'Муфта П1 530', 'full_name': 'Муфта КМТ 530', 'marker': 'шт', 'TU': 'ТУ8',
         'param': 'диаметр= мм, длина= мм, толщина стенки= м', 'date': '05.12.2019', 'doc': 'Паспорт 12', 'kol': '3534',
         'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'),
         'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'), 'kk': ('', '', ''), 'sk': ('', '', '')},
        {'name': 'Герметик', 'full_name': 'состав полмерный Герметик', 'marker': 'компл.', 'TU': 'ТУ6', 'param': '-',
         'date': '07.12.2019', 'doc': '645', 'kol': '3', 'otv': ('Кулешов А.Б.', 'Начальник РУ№3', 'ЦРС "Рязань"'),
         'contr': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'),
         'kk': ('Клименко Д.А.', 'Начальник ЛАЭС №1', 'ППС "Плавск'),
         'sk': ('Мирошкин М.В.', 'Инженер СК', 'ООО "Сег"')}]
    insert_akt_vh(akt_vh, road_programm, road_db, tube, km_start, km_finish, dy_tube)
